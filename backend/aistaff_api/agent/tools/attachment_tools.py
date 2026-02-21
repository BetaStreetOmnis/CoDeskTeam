from __future__ import annotations

import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from .base import ToolContext, ToolDefinition


_FILE_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,199}$")
_TEXT_EXTS = {".txt", ".md", ".markdown", ".csv", ".json", ".yaml", ".yml", ".log"}
_XLSX_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_MAX_BYTES_TEXT_READ = 2 * 1024 * 1024


def _safe_outputs_path(outputs_dir: Path, file_id: str) -> Path:
    if not _FILE_ID_RE.match(file_id) or ".." in file_id:
        raise ValueError("invalid file_id")

    base = outputs_dir.resolve()
    full = (base / file_id).resolve()
    if full == base or not str(full).startswith(str(base) + os.sep):
        raise ValueError("invalid file_id")
    return full


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    try:
        return str(value)
    except Exception:
        return ""


class AttachmentReadArgs(BaseModel):
    file_id: str = Field(min_length=1, max_length=200)
    mode: str = Field(default="auto", pattern="^(auto|text|xlsx)$")
    sheet: str | None = Field(default=None, max_length=200)
    max_sheets: int = Field(default=2, ge=1, le=6)
    max_rows: int = Field(default=25, ge=1, le=200)
    max_cols: int = Field(default=12, ge=1, le=50)


def attachment_read_tool() -> ToolDefinition:
    async def handler(args: AttachmentReadArgs, ctx: ToolContext) -> dict:
        full = _safe_outputs_path(ctx.outputs_dir, args.file_id)
        if not full.exists() or not full.is_file():
            raise ValueError("file not found")

        ext = full.suffix.lower()
        size_bytes = int(full.stat().st_size)

        mode = (args.mode or "auto").strip().lower()
        if mode == "auto":
            if ext in _XLSX_EXTS:
                mode = "xlsx"
            elif ext in _TEXT_EXTS:
                mode = "text"
            else:
                mode = "text" if size_bytes <= _MAX_BYTES_TEXT_READ else "meta"

        if mode == "meta":
            return {"file_id": args.file_id, "ext": ext, "size_bytes": size_bytes, "note": "binary file; preview not supported"}

        if mode == "text":
            data = full.read_bytes()
            if not data:
                return {"file_id": args.file_id, "ext": ext, "size_bytes": size_bytes, "text": ""}
            if len(data) > _MAX_BYTES_TEXT_READ:
                data = data[:_MAX_BYTES_TEXT_READ]
            text = data.decode("utf-8", errors="replace")
            if ctx.max_file_read_chars > 0 and len(text) > int(ctx.max_file_read_chars):
                text = text[: max(0, int(ctx.max_file_read_chars) - 14)] + "\n…(truncated)"
            return {"file_id": args.file_id, "ext": ext, "size_bytes": size_bytes, "text": text}

        if mode != "xlsx":
            raise ValueError("unsupported mode")

        wb = load_workbook(filename=str(full), read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames or [])
            selected: list[str] = []
            if args.sheet:
                if args.sheet not in sheet_names:
                    raise ValueError(f"sheet not found: {args.sheet}")
                selected = [args.sheet]
            else:
                selected = sheet_names[: int(args.max_sheets)]

            sheets_out: list[dict[str, Any]] = []
            for name in selected:
                ws = wb[name]
                max_row = int(getattr(ws, "max_row", 0) or 0)
                max_col = int(getattr(ws, "max_column", 0) or 0)
                rows: list[list[str]] = []
                rmax = min(max_row, int(args.max_rows))
                cmax = min(max_col, int(args.max_cols))
                if rmax > 0 and cmax > 0:
                    for row in ws.iter_rows(min_row=1, max_row=rmax, min_col=1, max_col=cmax, values_only=True):
                        rows.append([_cell_to_string(v) for v in row])
                sheets_out.append(
                    {
                        "name": name,
                        "max_row": max_row,
                        "max_col": max_col,
                        "preview_rows": rows,
                    }
                )

            return {
                "file_id": args.file_id,
                "ext": ext,
                "size_bytes": size_bytes,
                "sheet_names": sheet_names,
                "sheets": sheets_out,
                "note": "xlsx preview is truncated; use max_rows/max_cols/max_sheets to adjust",
            }
        finally:
            try:
                wb.close()
            except Exception:
                pass

    return ToolDefinition(
        name="attachment_read",
        description="读取用户上传/生成的附件内容（支持文本与 XLSX 预览）",
        risk="safe",
        input_model=AttachmentReadArgs,
        handler=handler,
        parameters_schema={
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "file_id": {"type": "string"},
                "mode": {"type": "string", "enum": ["auto", "text", "xlsx"], "default": "auto"},
                "sheet": {"type": "string"},
                "max_sheets": {"type": "integer", "minimum": 1, "maximum": 6, "default": 2},
                "max_rows": {"type": "integer", "minimum": 1, "maximum": 200, "default": 25},
                "max_cols": {"type": "integer", "minimum": 1, "maximum": 50, "default": 12},
            },
            "required": ["file_id"],
        },
    )

